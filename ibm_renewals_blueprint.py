"""
IBM S&S Renewal Module - Ingram Micro
Flask Blueprint — se integra al servidor Flask existente con 2 líneas.

INTEGRACIÓN:
    from ibm_renewals_blueprint import ibm_renewals_bp
    app.register_blueprint(ibm_renewals_bp, url_prefix='/renewals')
"""

from flask import Blueprint, request, jsonify, send_file
import os
import json
import zipfile
from datetime import datetime, date
from collections import defaultdict
import re
import base64
import openpyxl as _openpyxl

from xml_parser import parse_renewal_xml
from planilla_generator import check_coverage_errors, generate_planilla
from excel_generator import generate_internal_excel
from pdf_generator import generate_quote_pdf
from ui import register_ui_route
from rates_config import get_rates

ibm_renewals_bp = Blueprint('ibm_renewals', __name__)
register_ui_route(ibm_renewals_bp)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_file_from_request():
    if 'file' not in request.files:
        return None, jsonify({"success": False, "error": "No se envió ningún archivo"}), 400
    f = request.files['file']
    if not f.filename.endswith('.xml'):
        return None, jsonify({"success": False, "error": "El archivo debe ser XML (.xml)"}), 400
    return f, None, None

def _group_by_quote(renewals):
    groups = defaultdict(list)
    for item in renewals:
        quote_num = item.get('quote_number', 'SIN-QUOTE').strip() or 'SIN-QUOTE'
        groups[quote_num].append(item)
    return dict(groups)

def _is_ontime(renewal_due_date_str):
    try:
        for fmt in ('%d %b %Y', '%Y-%m-%d', '%m/%d/%Y'):
            try:
                due = datetime.strptime(renewal_due_date_str.strip(), fmt).date()
                return due >= date.today()
            except ValueError:
                continue
    except Exception:
        pass
    return False

def _apply_margins(renewals, bp_margin, ingram_margin, cvr_renew_bp, cvr_renew_ingram,
                   cvr_ontime_bp, cvr_ontime_ingram):
    for item in renewals:
        mep = float(item.get('extended_price', 0))
        qty = float(item.get('quantity', 1)) or 1
        unit_mep = float(item.get('unit_price', 0)) or round(mep / qty, 4)
        ontime = _is_ontime(str(item.get('renewal_due_date', '')))

        costo_canal          = round(mep * (1 - bp_margin / 100), 2)
        cvr_renew_bp_amt     = round(mep * (cvr_renew_bp / 100), 2)
        cvr_renew_ingram_amt = round(mep * (cvr_renew_ingram / 100), 2)
        cvr_ontime_bp_amt    = round(mep * (cvr_ontime_bp / 100), 2) if ontime else 0.0
        cvr_ontime_ingram_amt= round(mep * (cvr_ontime_ingram / 100), 2) if ontime else 0.0
        precio_final_bp      = round(costo_canal - cvr_renew_bp_amt - cvr_ontime_bp_amt, 2)
        costo_ingram         = round(costo_canal * (1 - ingram_margin / 100)
                                     - cvr_renew_ingram_amt - cvr_ontime_ingram_amt, 2)

        item.update({
            'mep': mep, 'unit_mep': unit_mep,
            'costo_canal': costo_canal,
            'cvr_renew_bp_amt': cvr_renew_bp_amt,
            'cvr_ontime_bp_amt': cvr_ontime_bp_amt,
            'cvr_renew_ingram_amt': cvr_renew_ingram_amt,
            'cvr_ontime_ingram_amt': cvr_ontime_ingram_amt,
            'precio_final_bp': precio_final_bp,
            'costo_ingram': costo_ingram,
            'ontime': ontime,
            'final_price': precio_final_bp,
            'bp_margin': bp_margin,
            'ingram_margin': ingram_margin,
            'cvr_renew_bp': cvr_renew_bp,
            'cvr_ontime_bp': cvr_ontime_bp,
        })
    return renewals

def _safe_filename(text):
    return re.sub(r'[^\w\-. ]', '_', str(text)).strip()

def _extract_reseller_name(reseller_str):
    if not reseller_str:
        return 'SIN_RESELLER'
    match = re.match(r'^\d+\s+(.+)', str(reseller_str))
    return match.group(1).strip() if match else str(reseller_str).strip()

def _extract_customer_short(site_str):
    if not site_str:
        return 'SIN_CLIENTE'
    match = re.match(r'^\d+-(.+?)(?:,|$)', str(site_str))
    return match.group(1).strip() if match else str(site_str).strip()

def _build_quote_files(lines, quote_num, bp_margin, ingram_margin,
                       cvr_renew_bp, cvr_ontime_bp):
    year          = datetime.now().year
    reseller_name = _extract_reseller_name(lines[0].get('reseller', ''))
    customer_short= _extract_customer_short(lines[0].get('site', ''))
    customer_full = lines[0].get('site', 'Cliente')
    base_name     = f"S&S - {customer_short} - {reseller_name} - {year}"

    pdf_path = os.path.join(OUTPUT_DIR, f"{_safe_filename(base_name)}.pdf")
    generate_quote_pdf(lines, customer_full,
                       f"Renovación S&S IBM — {customer_short}",
                       bp_margin, pdf_path,
                       quote_number=quote_num,
                       cvr_renew_bp=cvr_renew_bp,
                       cvr_ontime_bp=cvr_ontime_bp)

    excel_path = os.path.join(OUTPUT_DIR, f"{_safe_filename(base_name)}.xlsx")
    generate_internal_excel(lines, quote_num, excel_path)

    planilla_path = None
    if check_coverage_errors(lines):
        planilla_name = (f"Planilla de renovación SSA y MX - "
                         f"{customer_short} - {reseller_name} - 12 Meses - {year}")
        planilla_path = os.path.join(OUTPUT_DIR, f"{_safe_filename(planilla_name)}.xlsx")
        generate_planilla(lines, quote_num, planilla_path)

    return reseller_name, customer_short, pdf_path, excel_path, planilla_path

# ── Endpoints ─────────────────────────────────────────────────────────────────

@ibm_renewals_bp.get('/health')
def health():
    r = get_rates()
    return jsonify({
        "status":      "ok",
        "module":      "IBM S&S Renewals",
        "timestamp":   datetime.now().isoformat(),
        "rates_label": r['label'],
        "rates": {k: v for k, v in r.items()
                  if k not in ('label', 'valid_from', 'valid_until', 'rate_key')},
    })

@ibm_renewals_bp.get('/rates')
def get_current_rates():
    r = get_rates()
    return jsonify({
        "success":           True,
        "label":             r['label'],
        "rate_key":          r['rate_key'],
        "bp_margin":         r['bp_margin'],
        "ingram_margin":     r['ingram_margin'],
        "cvr_renew_bp":      r['cvr_renew_bp'],
        "cvr_renew_ingram":  r['cvr_renew_ingram'],
        "cvr_ontime_bp":     r['cvr_ontime_bp'],
        "cvr_ontime_ingram": r['cvr_ontime_ingram'],
    })

@ibm_renewals_bp.post('/parse-renewal-report')
def parse_renewal_report():
    f, err, code = _get_file_from_request()
    if err:
        return err, code
    try:
        renewals = parse_renewal_xml(f.read())
        groups   = _group_by_quote(renewals)
        quotes   = {}
        for qn, lines in groups.items():
            quotes[qn] = {
                "lines":      lines,
                "customer":   lines[0].get('site', 'Cliente'),
                "total_ibm":  sum(float(l.get('extended_price', 0)) for l in lines),
                "line_count": len(lines),
                "ontime":     _is_ontime(str(lines[0].get('renewal_due_date', ''))),
            }
        coverage_alerts = {}
        for qn, qdata in quotes.items():
            errors = check_coverage_errors(qdata['lines'])
            if errors:
                coverage_alerts[qn] = [{
                    'part_number': e['part_number'],
                    'months':      e['months'],
                    'start':       e['start_str'],
                    'end':         e['end_str'],
                    'correct_end': e['correct_end'],
                } for e in errors]
        return jsonify({"success": True, "total_lines": len(renewals),
                        "total_quotes": len(groups), "quotes": quotes,
                        "coverage_alerts": coverage_alerts})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@ibm_renewals_bp.post('/generate-planilla')
def generate_planilla_endpoint():
    f, err, code = _get_file_from_request()
    if err:
        return err, code
    quote_filter = request.form.get('quote_filter', '').strip() or None
    try:
        renewals = parse_renewal_xml(f.read())
        groups   = _group_by_quote(renewals)
        lines    = groups.get(quote_filter, renewals) if quote_filter else renewals
        timestamp= datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(OUTPUT_DIR,
                        f"Planilla_IBM_{_safe_filename(quote_filter or 'all')}_{timestamp}.xlsx")
        result = generate_planilla(lines, quote_filter or 'all', output_path)
        if not result:
            return jsonify({"success": False,
                            "error": "No hay errores de coverage en este quote"}), 400
        return send_file(output_path,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=os.path.basename(output_path))
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@ibm_renewals_bp.post('/generate-quote')
def generate_quote():
    f, err, code = _get_file_from_request()
    if err:
        return err, code
    try:
        _r = get_rates()
        bp_margin        = float(request.form.get('bp_margin',        _r['bp_margin']))
        ingram_margin    = float(request.form.get('ingram_margin',    _r['ingram_margin']))
        cvr_renew_bp     = float(request.form.get('cvr_renew_bp',     _r['cvr_renew_bp']))
        cvr_renew_ingram = float(request.form.get('cvr_renew_ingram', _r['cvr_renew_ingram']))
        cvr_ontime_bp    = float(request.form.get('cvr_ontime_bp',    _r['cvr_ontime_bp']))
        cvr_ontime_ingram= float(request.form.get('cvr_ontime_ingram',_r['cvr_ontime_ingram']))
    except (ValueError, TypeError):
        return jsonify({"success": False, "error": "Los márgenes deben ser números"}), 400

    quote_filter = request.form.get('quote_filter', '').strip() or None
    try:
        renewals = parse_renewal_xml(f.read())
        if not renewals:
            return jsonify({"success": False,
                            "error": "No se encontraron líneas de renovación"}), 400
        renewals = _apply_margins(renewals, bp_margin, ingram_margin,
                                  cvr_renew_bp, cvr_renew_ingram,
                                  cvr_ontime_bp, cvr_ontime_ingram)
        groups = _group_by_quote(renewals)
        if quote_filter and quote_filter in groups:
            groups = {quote_filter: groups[quote_filter]}

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        year      = datetime.now().year
        entries   = []
        for quote_num, lines in groups.items():
            reseller_name, customer_short, pdf_p, excel_p, planilla_p = \
                _build_quote_files(lines, quote_num, bp_margin, ingram_margin,
                                   cvr_renew_bp, cvr_ontime_bp)
            entries.append((reseller_name, customer_short, pdf_p, excel_p, planilla_p))

        q0       = list(groups.keys())[0]
        zip_name = (f"Quote_{_safe_filename(q0)}_{timestamp}.zip"
                    if len(entries) == 1
                    else f"Cotizaciones_IBM_SS_{year}_{timestamp}.zip")
        zip_path = os.path.join(OUTPUT_DIR, zip_name)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for reseller, customer_short, pdf_p, excel_p, planilla_p in entries:
                folder = f"{year}/{reseller}/"
                zf.write(pdf_p,   folder + os.path.basename(pdf_p))
                zf.write(excel_p, folder + os.path.basename(excel_p))
                if planilla_p:
                    zf.write(planilla_p, folder + os.path.basename(planilla_p))

        return send_file(zip_path, mimetype='application/zip',
                         as_attachment=True, download_name=zip_name)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@ibm_renewals_bp.post('/generate-bulk')
def generate_bulk():
    f, err, code = _get_file_from_request()
    if err:
        return err, code

    try:
        _r = get_rates()
        bp_margin        = float(request.form.get('bp_margin',        _r['bp_margin']))
        ingram_margin    = float(request.form.get('ingram_margin',    _r['ingram_margin']))
        cvr_renew_bp     = float(request.form.get('cvr_renew_bp',     _r['cvr_renew_bp']))
        cvr_renew_ingram = float(request.form.get('cvr_renew_ingram', _r['cvr_renew_ingram']))
        cvr_ontime_bp    = float(request.form.get('cvr_ontime_bp',    _r['cvr_ontime_bp']))
        cvr_ontime_ingram= float(request.form.get('cvr_ontime_ingram',_r['cvr_ontime_ingram']))
    except (ValueError, TypeError):
        return jsonify({"success": False, "error": "Márgenes inválidos"}), 400

    try:
        quote_numbers = json.loads(request.form.get('quote_numbers', '[]'))
        quote_params  = json.loads(request.form.get('quote_params',  '{}'))
    except Exception:
        return jsonify({"success": False,
                        "error": "Formato inválido en quote_numbers o quote_params"}), 400

    if not quote_numbers:
        return jsonify({"success": False,
                        "error": "No se enviaron quotes para generar"}), 400

    try:
        renewals_all = parse_renewal_xml(f.read())
        groups       = _group_by_quote(renewals_all)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    year      = datetime.now().year
    zip_name  = f"IBM_SS_Renovaciones_{timestamp}_{len(quote_numbers)}quotes.zip"
    zip_path  = os.path.join(OUTPUT_DIR, zip_name)

    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for qn in quote_numbers:
                if qn not in groups:
                    continue
                lines = list(groups[qn])
                p        = quote_params.get(str(qn), {})
                q_bp     = float(p.get('bp',          bp_margin))
                q_ingram = float(p.get('ingram',       ingram_margin))
                q_cvr_r  = float(p.get('cvrRenew',     cvr_renew_bp))
                q_cvr_ot = float(p.get('cvrOntime',    cvr_ontime_bp))
                q_cvr_ri = float(p.get('cvrRenewIng',  cvr_renew_ingram))
                q_cvr_oi = float(p.get('cvrOntimeIng', cvr_ontime_ingram))

                if p.get('useRenew') is False:
                    q_cvr_r = q_cvr_ri = 0.0
                if p.get('useOntime') is False:
                    q_cvr_ot = q_cvr_oi = 0.0

                lines = _apply_margins(lines, q_bp, q_ingram,
                                       q_cvr_r, q_cvr_ri, q_cvr_ot, q_cvr_oi)
                reseller_name, customer_short, pdf_p, excel_p, planilla_p = \
                    _build_quote_files(lines, qn, q_bp, q_ingram, q_cvr_r, q_cvr_ot)

                folder = f"{_safe_filename(reseller_name)}/{year}/"
                zf.write(pdf_p,   folder + os.path.basename(pdf_p))
                zf.write(excel_p, folder + os.path.basename(excel_p))
                if planilla_p:
                    zf.write(planilla_p, folder + os.path.basename(planilla_p))

        return send_file(zip_path, mimetype='application/zip',
                         as_attachment=True, download_name=zip_name)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    
# ── AGREGAR ESTE ENDPOINT AL FINAL DE ibm_renewals_blueprint.py ──────────────
# Requiere: import base64, import openpyxl (agregar al inicio del archivo)

import base64
import openpyxl as _openpyxl

@ibm_renewals_bp.post('/generate-preview')
def generate_preview():
    """
    Genera PDF + Excel + Planilla para un quote y los devuelve como base64
    para embeber en la UI sin necesidad de descarga.
    """
    f, err, code = _get_file_from_request()
    if err:
        return err, code

    try:
        _r = get_rates()
        bp_margin        = float(request.form.get('bp_margin',        _r['bp_margin']))
        ingram_margin    = float(request.form.get('ingram_margin',    _r['ingram_margin']))
        cvr_renew_bp     = float(request.form.get('cvr_renew_bp',     _r['cvr_renew_bp']))
        cvr_renew_ingram = float(request.form.get('cvr_renew_ingram', _r['cvr_renew_ingram']))
        cvr_ontime_bp    = float(request.form.get('cvr_ontime_bp',    _r['cvr_ontime_bp']))
        cvr_ontime_ingram= float(request.form.get('cvr_ontime_ingram',_r['cvr_ontime_ingram']))
    except (ValueError, TypeError):
        return jsonify({"success": False, "error": "Márgenes inválidos"}), 400

    quote_filter = request.form.get('quote_filter', '').strip() or None

    try:
        renewals = parse_renewal_xml(f.read())
        if not renewals:
            return jsonify({"success": False, "error": "No se encontraron líneas"}), 400

        renewals = _apply_margins(renewals, bp_margin, ingram_margin,
                                  cvr_renew_bp, cvr_renew_ingram,
                                  cvr_ontime_bp, cvr_ontime_ingram)
        groups = _group_by_quote(renewals)

        if quote_filter and quote_filter in groups:
            lines = groups[quote_filter]
        else:
            return jsonify({"success": False, "error": "Quote no encontrado"}), 404

        year          = datetime.now().year
        reseller_name = _extract_reseller_name(lines[0].get('reseller', ''))
        customer_short= _extract_customer_short(lines[0].get('site', ''))
        customer_full = lines[0].get('site', 'Cliente')
        base_name     = f"S&S - {customer_short} - {reseller_name} - {year}"
        result        = {}

        # PDF
        pdf_path = os.path.join(OUTPUT_DIR, f"preview_{_safe_filename(quote_filter)}.pdf")
        generate_quote_pdf(lines, customer_full,
                           f"Renovación S&S IBM — {customer_short}",
                           bp_margin, pdf_path,
                           quote_number=quote_filter,
                           cvr_renew_bp=cvr_renew_bp,
                           cvr_ontime_bp=cvr_ontime_bp)
        with open(pdf_path, 'rb') as fp:
            result['pdf_b64'] = base64.b64encode(fp.read()).decode('utf-8')
        result['pdf_filename'] = f"{_safe_filename(base_name)}.pdf"

        # Excel — leer hojas como JSON
        excel_path = os.path.join(OUTPUT_DIR, f"preview_{_safe_filename(quote_filter)}.xlsx")
        generate_internal_excel(lines, quote_filter, excel_path)
        wb = _openpyxl.load_workbook(excel_path, data_only=True)
        sheets = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True):
                rows.append([str(c) if c is not None else '' for c in row])
            sheets[sn] = rows
        result['excel_sheets']   = sheets
        result['excel_filename'] = f"{_safe_filename(base_name)}.xlsx"

        # Planilla (solo si hay coverage incorrecto)
        coverage_errors = check_coverage_errors(lines)
        if coverage_errors:
            planilla_name = (f"Planilla de renovación SSA y MX - "
                             f"{customer_short} - {reseller_name} - 12 Meses - {year}")
            planilla_path = os.path.join(OUTPUT_DIR, f"preview_planilla_{_safe_filename(quote_filter)}.xlsx")
            generate_planilla(lines, quote_filter, planilla_path)
            wb2 = _openpyxl.load_workbook(planilla_path, data_only=True)
            sheets2 = {}
            for sn in wb2.sheetnames:
                ws2 = wb2[sn]
                rows2 = []
                for row in ws2.iter_rows(min_row=1, max_row=min(ws2.max_row, 60), values_only=True):
                    rows2.append([str(c) if c is not None else '' for c in row])
                sheets2[sn] = rows2
            result['planilla_sheets']   = sheets2
            result['planilla_filename'] = f"{_safe_filename(planilla_name)}.xlsx"

        result['success'] = True
        return jsonify(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
