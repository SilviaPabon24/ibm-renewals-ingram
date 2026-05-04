"""
Generador de Planilla de Renovación IBM (SSA/MX)
Llena la plantilla base con los datos del quote cuando el coverage es incorrecto.
"""

import openpyxl
from openpyxl import load_workbook
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import os
import re
import shutil

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'plantilla_base.xlsx')


def _parse_date(date_str):
    """Parsea fechas en formatos comunes del reporte IBM."""
    if not date_str:
        return None
    for fmt in ('%d %b %Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(str(date_str).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _correct_end_date(start_date):
    """Calcula la fecha fin correcta: inicio + 12 meses - 1 día."""
    return start_date + relativedelta(months=12) - relativedelta(days=1)


def _coverage_months(start_str, end_str):
    """Retorna los meses de cobertura entre dos fechas.
    Usa end + 1 día para manejar correctamente rangos como 01 Apr → 31 Mar (= 12 meses).
    """
    from datetime import timedelta
    start = _parse_date(start_str)
    end   = _parse_date(end_str)
    if not start or not end:
        return None
    delta = relativedelta(end + timedelta(days=1), start)
    return delta.months + delta.years * 12


def check_coverage_errors(renewals):
    """
    Revisa si alguna línea tiene coverage diferente a 12 meses.
    Retorna lista de dicts con las líneas con error.
    Siempre usa coverage_dates para extraer inicio y fin reales.
    """
    errors = []
    for item in renewals:
        cov = item.get('coverage_dates', '')
        if ' To ' in cov:
            start_str = cov.split(' To ')[0].strip()
            end_str   = cov.split(' To ')[-1].strip()
        else:
            start_str = item.get('renewal_line_item_start_date', '')
            end_str   = item.get('renewal_due_date', '')
        months    = _coverage_months(start_str, end_str)
        if months is not None and months != 12:
            start = _parse_date(start_str)
            correct_end = _correct_end_date(start) if start else None
            errors.append({
                'part_number':   item.get('part_number', ''),
                'part_description': item.get('part_description', ''),
                'quantity':      item.get('quantity', 1),
                'start_str':     start_str,
                'end_str':       end_str,
                'months':        months,
                'correct_end':   correct_end.strftime('%d/%m/%Y') if correct_end else '',
                'start_fmt':     start.strftime('%d/%m/%Y') if start else start_str,
            })
    return errors


def _extract_reseller_site(reseller_str):
    """Extrae el site number del string 'XXXXXXX Nombre Reseller'."""
    if not reseller_str:
        return ''
    match = re.match(r'^(\d+)\s+', str(reseller_str))
    return match.group(1) if match else ''


def _extract_reseller_name(reseller_str):
    """Extrae el nombre del reseller del string 'XXXXXXX Nombre Reseller'."""
    if not reseller_str:
        return ''
    match = re.match(r'^\d+\s+(.+)', str(reseller_str))
    return match.group(1).strip() if match else str(reseller_str)


def _extract_customer_name(site_str):
    """Extrae el nombre del cliente del string 'XXXXXXX-Nombre Cliente, Dirección, COL'."""
    if not site_str:
        return ''
    match = re.match(r'^\d+-(.+?)(?:,|$)', str(site_str))
    return match.group(1).strip() if match else str(site_str)


def _extract_agreement(agreement_str):
    """Extrae el número de acuerdo del string 'XXXXXXXXXX-STD'."""
    if not agreement_str or agreement_str == 'null-null':
        return ''
    match = re.match(r'^(\d+)', str(agreement_str))
    return match.group(1) if match else ''


def generate_planilla(lines, quote_number, output_path):
    """
    Genera la planilla de corrección IBM para un quote con coverage incorrecto.
    Solo llena los campos requeridos en Factsheet y SBid.
    """
    errors = check_coverage_errors(lines)
    if not errors:
        return None  # No hay errores de coverage, no se genera plantilla

    # Cargar plantilla base
    wb = load_workbook(TEMPLATE_PATH)

    # ── HOJA FACTSHEET ────────────────────────────────────────────────────────
    fs = wb['Factsheet']

    first = lines[0]
    customer_name   = _extract_customer_name(first.get('site', ''))
    customer_site   = first.get('ibm_customer_number', '')
    agreement       = _extract_agreement(first.get('agreement_information', ''))
    reseller_name   = _extract_reseller_name(first.get('reseller', ''))
    reseller_site   = _extract_reseller_site(first.get('reseller', ''))

    # Country (B5 = row 5, col B)
    fs['B5'] = 'Colombia'

    # Customer Name (B7), Customer Site Number (E7), Agreement (I7)
    fs['B7'] = customer_name
    fs['E7'] = customer_site
    fs['I7'] = agreement

    # Distributor (B10), Reseller Name (E10), Reseller Site (I10)
    fs['B10'] = 'Ingram Micro Colombia'
    fs['E10'] = reseller_name
    fs['I10'] = reseller_site

    # RTD (B13 ya tiene 'Manual Bid', no tocar)

    # ── HOJA SBID ─────────────────────────────────────────────────────────────
    sb = wb['SBid ']

    # Llenar líneas con errores de coverage a partir de fila 26
    start_row = 26
    for i, err in enumerate(errors):
        row = start_row + i
        sb.cell(row=row, column=1).value  = err['part_description']   # Description
        sb.cell(row=row, column=2).value  = err['part_number']         # PN
        sb.cell(row=row, column=4).value  = err['quantity']            # Qty
        sb.cell(row=row, column=16).value = err['start_fmt']           # Begin Cvg Date
        sb.cell(row=row, column=17).value = err['correct_end']         # End Cvg Date (corregida)

    wb.save(output_path)
    return output_path
