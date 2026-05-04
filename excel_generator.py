"""
Generador de Excel — Renovaciones IBM S&S (Ingram Micro)
Replica la estructura de PLANTILLA_RNW.xlsx con 4 hojas:
  1. Info         — datos del negocio (inputs manuales + del XML)
  2. Calculadora  — motor de precios con fórmulas cruzadas desde Info
  3. Cotización   — propuesta para el canal, jalando de Calculadora
  4. Compra       — orden interna para el equipo de compras
"""

import re
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Paleta Ingram ─────────────────────────────────────────────────────────────
INGRAM_BLUE  = '0066CC'
INGRAM_DARK  = '003366'
LIGHT_BLUE   = 'EBF3FF'
HEADER_GRAY  = 'F1F5F9'
WHITE        = 'FFFFFF'
YELLOW_INPUT = 'FFF9C4'
GREEN_LINK   = 'E8F5E9'   # celdas que jalón de otra hoja (convención verde)
BLUE_INPUT   = 'DBEAFE'   # inputs hardcoded (convención azul)
GRAY_TEXT    = '64748B'
DARK_TEXT    = '1A2B3C'
ALERT_RED    = 'FFEBEE'

USD_FMT  = '"$"#,##0.00'
PCT_FMT  = '0.00%'
DATE_FMT = 'DD/MM/YYYY'


# ── Helpers de estilo ─────────────────────────────────────────────────────────
def _h(cell, bg=INGRAM_BLUE, fg=WHITE, bold=True, sz=9, wrap=False, align='center'):
    cell.font      = Font(name='Arial', bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)

def _v(cell, bold=False, color=DARK_TEXT, sz=9, align='left', bg=None, wrap=False):
    cell.font      = Font(name='Arial', bold=bold, color=color, size=sz)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if bg:
        cell.fill  = PatternFill('solid', fgColor=bg)

def _border(color='D1D5DB'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thin_bottom(color='D1D5DB'):
    s = Side(style='thin', color=color)
    return Border(bottom=s)

def _usd(cell):   cell.number_format = USD_FMT
def _pct(cell):   cell.number_format = PCT_FMT
def _date(cell):  cell.number_format = DATE_FMT


# ── Extracción de datos del XML ───────────────────────────────────────────────
def _reseller_name(s):
    m = re.match(r'^\d+\s+(.+)', str(s or ''))
    return m.group(1).strip() if m else str(s or '')

def _reseller_site(s):
    m = re.match(r'^(\d+)\s+', str(s or ''))
    return m.group(1) if m else ''

def _customer_name(s):
    m = re.match(r'^\d+-(.+?)(?:,|$)', str(s or ''))
    return m.group(1).strip() if m else str(s or '')

def _customer_site(s):
    m = re.match(r'^(\d+)-', str(s or ''))
    return m.group(1) if m else ''

def _agreement(s):
    if not s or s == 'null-null':
        return ''
    m = re.match(r'^(\d+)', str(s))
    return m.group(1) if m else ''

def _parse_cov_date(cov_str, part):
    """Extrae fecha inicio (part=0) o fin (part=1) de 'DD Mon YYYY To DD Mon YYYY'."""
    if ' To ' in str(cov_str):
        return str(cov_str).split(' To ')[part].strip()
    return ''


# ── HOJA 1: INFO ──────────────────────────────────────────────────────────────
def _build_info(ws, lines, quote_number, bp_margin, ingram_margin,
                cvr_renew_bp, cvr_renew_ingram, cvr_ontime_bp, cvr_ontime_ingram):
    first = lines[0]

    # Anchos
    for col, w in [('A',3),('B',3),('C',18),('D',28),('E',3),
                   ('F',18),('G',18),('H',3),('I',18),('J',14)]:
        ws.column_dimensions[col].width = w

    # ── Título ──────────────────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    ws['A1'] = 'INGRAM MICRO — IBM S&S Renewals · Hoja de Información'
    _h(ws['A1'], bg=INGRAM_DARK, sz=11)
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Quote: {quote_number}  ·  Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    _v(ws['A2'], color=GRAY_TEXT, sz=8, align='center')
    ws.row_dimensions[2].height = 14

    def sec(row, txt):
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = txt
        _h(ws[f'A{row}'], bg=INGRAM_BLUE, sz=9, align='left')
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 18

    def lbl(row, col_l, col_v, label, value, bg_v=None, is_pct=False, is_date=False, color_v=DARK_TEXT):
        cl = ws.cell(row=row, column=col_l, value=label)
        _v(cl, bold=True, color=GRAY_TEXT, sz=9)
        cv = ws.cell(row=row, column=col_v, value=value)
        _v(cv, sz=9, bg=bg_v, color=color_v)
        cv.border = _border()
        if is_pct:   _pct(cv)
        if is_date:  _date(cv)
        ws.row_dimensions[row].height = 16
        return cv

    # ── Sección CANAL (filas 4-7) ────────────────────────────────────────────
    sec(3, '  CANAL')
    lbl(4,  3, 4, 'CANAL',          _reseller_name(first.get('reseller','')),  bg_v=BLUE_INPUT)
    lbl(5,  3, 4, 'CLIENTE',        _customer_name(first.get('site','')),       bg_v=BLUE_INPUT)
    lbl(6,  3, 4, 'CÓDIGO CANAL',   _reseller_site(first.get('reseller','')),   bg_v=BLUE_INPUT)
    lbl(4,  6, 7, 'ICN',            first.get('ibm_customer_number',''),         bg_v=BLUE_INPUT)
    lbl(5,  6, 7, 'SITE',           first.get('site',''),                        bg_v=BLUE_INPUT)
    lbl(6,  6, 7, 'Select Territory', 'Yes' if first.get('select_territory_accelerator','') else 'No', bg_v=BLUE_INPUT)
    lbl(4,  9, 10,'SBID/QUOTE',     quote_number,                                bg_v=YELLOW_INPUT)

    # ── Sección CLASIFICACIÓN (filas 9-13) ───────────────────────────────────
    sec(8, '  CLASIFICACIÓN')
    lbl(9,  3, 4, 'Cliente',        _customer_name(first.get('site','')),       bg_v=BLUE_INPUT)
    lbl(10, 3, 4, 'Banda',          'BL',                                        bg_v=BLUE_INPUT)
    lbl(11, 3, 4, 'Gobierno',       'Yes' if 'YES' in str(first.get('eligible_government_program','')).upper() else 'No', bg_v=BLUE_INPUT)
    lbl(12, 3, 4, 'Categoría',      'S&S',                                       bg_v=BLUE_INPUT)
    lbl(13, 3, 4, 'Cliente',        first.get('customer_set_designation','CO-LED'), bg_v=BLUE_INPUT)

    # ── Sección INCENTIVOS / MÁRGENES (filas 9-14, cols F-J) ─────────────────
    # Headers incentivos
    ws['F9']  = 'Incentivos';  _h(ws['F9'],  bg=HEADER_GRAY, fg=DARK_TEXT, sz=9)
    ws['F10'] = 'BP';          _h(ws['F10'], bg=HEADER_GRAY, fg=DARK_TEXT, sz=9)
    ws['G10'] = 'INGRAM';      _h(ws['G10'], bg=HEADER_GRAY, fg=DARK_TEXT, sz=9)

    ws['F11'] = 'CVR Renew'
    _v(ws['F11'], bold=True, sz=9)
    # F12 = CVR Renew BP  (input — celda que el usuario llenaba manualmente, ahora precargada)
    ws['F12'] = cvr_renew_bp / 100
    _v(ws['F12'], bg=YELLOW_INPUT, sz=9)
    ws['F12'].border = _border()
    _pct(ws['F12'])
    ws['G12'] = cvr_renew_ingram / 100
    _v(ws['G12'], bg=YELLOW_INPUT, sz=9)
    ws['G12'].border = _border()
    _pct(ws['G12'])

    ws['F13'] = 'CVR OnTime'
    _v(ws['F13'], bold=True, sz=9)
    ws['F14'] = cvr_ontime_bp / 100
    _v(ws['F14'], bg=YELLOW_INPUT, sz=9)
    ws['F14'].border = _border()
    _pct(ws['F14'])
    ws['G14'] = cvr_ontime_ingram / 100
    _v(ws['G14'], bg=YELLOW_INPUT, sz=9)
    ws['G14'].border = _border()
    _pct(ws['G14'])

    # Márgenes cadena (I9:J11)
    ws['I9']  = 'Margen Cadena';  _h(ws['I9'],  bg=HEADER_GRAY, fg=DARK_TEXT, sz=9)
    ws['I10'] = 'Margen Canal';   _v(ws['I10'], bold=True, sz=9)
    ws['J10'] = bp_margin / 100
    _v(ws['J10'], bg=YELLOW_INPUT, sz=9)
    ws['J10'].border = _border()
    _pct(ws['J10'])

    ws['I11'] = 'Margen Ingram';  _v(ws['I11'], bold=True, sz=9)
    ws['J11'] = ingram_margin / 100
    _v(ws['J11'], bg=YELLOW_INPUT, sz=9)
    ws['J11'].border = _border()
    _pct(ws['J11'])

    # J9 = suma J10+J11
    ws['J9'] = '=J10+J11'
    _v(ws['J9'], bold=True, color='000000', sz=9)
    ws['J9'].border = _border()
    _pct(ws['J9'])

    for r in range(9, 15):
        ws.row_dimensions[r].height = 16


# ── HOJA 2: CALCULADORA ───────────────────────────────────────────────────────
def _build_calculadora(ws, lines):
    """
    Replica el motor de cálculo de la plantilla original.
    Columnas clave (igual que original):
      B=IngramOnTime, C=IngramRenew, D=BPOnTime, E=BPRenew
      F=FechaInicio, G=FechaFin, H=NP, I=Descripción
      J=PrecioUnitBandaSinUplift, K=PrecioUnitBandaConUplift
      L=Cantidad, M=PrecioBanda, N=Descuento, O=PrecioClienteFinal
      P=MargenCanal, Q=V/unitVenta, R=CostoCanal
      S=CostoIngramUnit, T=PrecioCompraIngram, U=CostoIngram, V=MargenIngramTX
      W=Año, X=PrecioUnitClienteFinal
      Z=%IncentivoTotal, AA=Incentivo, AB=%BP, AC=TotalIncentivos
      AD=OnTime, AE=Renew, AF=IncTotalUnit, AG=IncBPUnit, AH=IncIngramUnit
      AJ=CostoTotalBP, AK=CostoIngramTotal, AL=MargenAntesInc, AM=MargenDespuesInc
    """

    # Anchos columnas
    widths = {
        'A':3,'B':8,'C':8,'D':8,'E':8,'F':12,'G':12,'H':16,'I':32,
        'J':12,'K':12,'L':8,'M':12,'N':8,'O':14,'P':10,'Q':12,'R':12,
        'S':14,'T':14,'U':12,'V':10,'W':8,'X':14,'Y':3,
        'Z':12,'AA':12,'AB':8,'AC':12,'AD':10,'AE':10,
        'AF':14,'AG':14,'AH':14,'AI':3,'AJ':14,'AK':14,'AL':12,'AM':12,
        'AN':3,'AO':3,'AP':8,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Título ──────────────────────────────────────────────────────────────
    ws.merge_cells('A1:AM1')
    ws['A1'] = 'CALCULADORA — IBM S&S Renewals · Ingram Micro'
    _h(ws['A1'], bg=INGRAM_DARK, sz=11)
    ws.row_dimensions[1].height = 24

    # ── Grupos de header (filas 3-4) ─────────────────────────────────────────
    # Fila 3: grupos macro
    groups_3 = [
        ('B3:E3', 'INGRAM / BP'),
        ('F3:G3', 'Fechas'),
        ('H3:K3', 'Producto'),
        ('L3:O3', 'Precios IBM'),
        ('P3:V3', 'Cadena de precios'),
        ('Z3:AH3', 'Incentivos Total — BP'),
        ('AJ3:AM3', 'Después de incentivos'),
    ]
    for rng, txt in groups_3:
        ws.merge_cells(rng)
        c = ws[rng.split(':')[0]]
        c.value = txt
        _h(c, bg=INGRAM_DARK, sz=8)
        ws.row_dimensions[3].height = 14

    # Fila 4: sub-headers INGRAM/BP
    for col, txt in [('B','On Time'),('C','Renew'),('D','On Time'),('E','Renew')]:
        c = ws[f'{col}4']
        c.value = txt
        _h(c, bg='475569', sz=8)
    ws.merge_cells('A3:A4'); ws['A3'].value = ''
    ws.row_dimensions[4].height = 14

    # ── Fila 5: headers columna por columna ──────────────────────────────────
    headers = {
        'B':'On Time','C':'Renew','D':'On Time','E':'Renew',
        'F':'Fecha de Inicio','G':'Fecha de Fin','H':'NP','I':'Descripción',
        'J':'Precio Unitario\nBanda sin Uplift','K':'Precio Unitario\nBanda con Uplift',
        'L':'Cantidad','M':'Precio Banda','N':'Descuento',
        'O':'Precio Cliente\nFinal','P':'Margen canal','Q':'V/unit Venta',
        'R':'Costo Canal','S':'Costo Ingram Unit','T':'Precio Compra Ingram',
        'U':'Costo Ingram','V':'Margen Ingram TX','W':'Año',
        'X':'Precio unitario\ncliente final',
        'Z':'% Incentivos Total','AA':'Incentivo','AB':'% BP',
        'AC':'Total Incentivos','AD':'On time','AE':'Renew',
        'AF':'Incentivo total\nUnitario','AG':'Incentivo BP\nUnitario',
        'AH':'Incentivo Ingram\nUnitario',
        'AJ':'Costo Total BP','AK':'Costo Ingram Total',
        'AL':'Margen antes\nde incentivos','AM':'Margen después\nde incentivos',
    }
    for col, txt in headers.items():
        c = ws[f'{col}5']
        c.value = txt
        _h(c, bg=INGRAM_BLUE, sz=8, wrap=True)
    ws.row_dimensions[5].height = 30

    # ── Filas de datos ────────────────────────────────────────────────────────
    first_data_row = 6
    n = len(lines)

    for i, item in enumerate(lines):
        r = first_data_row + i
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE

        cov = str(item.get('coverage_dates', ''))
        start_str = _parse_cov_date(cov, 0) or str(item.get('renewal_line_item_start_date',''))
        end_str   = _parse_cov_date(cov, 1) or str(item.get('renewal_due_date',''))

        qty   = int(item.get('quantity', 1))
        mep   = float(item.get('extended_price', 0))
        unit_mep = float(item.get('unit_price', 0)) or (round(mep / qty, 4) if qty else 0)

        # Ingram On/Renew (jalados de Info via fórmula)
        for col in ['B','C','D','E']:
            c = ws.cell(row=r, column=_col(col))
            c.value = f'=Info!${"G" if col in ["B","C"] else "F"}${"14" if col in ["B","D"] else "12"}'
            _v(c, sz=8, bg=GREEN_LINK, color='166534')
            _pct(c)

        # Fechas
        for col, val in [('F', start_str), ('G', end_str)]:
            c = ws.cell(row=r, column=_col(col))
            try:
                from datetime import datetime as dt
                for fmt in ('%d %b %Y','%Y-%m-%d','%m/%d/%Y','%d/%m/%Y'):
                    try:
                        c.value = dt.strptime(val.strip(), fmt).date()
                        _date(c)
                        break
                    except:
                        continue
                else:
                    c.value = val
            except:
                c.value = val
            _v(c, sz=8, bg=bg)

        # Producto
        _set(ws, r, 'H', item.get('part_number',''),       bg, sz=8)
        _set(ws, r, 'I', item.get('part_description',''),  bg, sz=8, wrap=True)
        _set(ws, r, 'J', unit_mep,  bg, sz=8, fmt=USD_FMT)
        _set(ws, r, 'K', unit_mep,  bg, sz=8, fmt=USD_FMT)  # con uplift = sin uplift en S&S
        _set(ws, r, 'L', qty,       bg, sz=8, align='center')
        _set(ws, r, 'W', 'Año 1',   bg, sz=8, align='center')

        # Precio Banda = Precio Unit * Cantidad
        c = ws.cell(row=r, column=_col('M'))
        c.value = f'=K{r}*L{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Descuento (0 para S&S, el MEP ya es el precio final)
        _set(ws, r, 'N', 0, bg, sz=8, fmt=PCT_FMT)

        # Precio Cliente Final = M*(1-N)
        c = ws.cell(row=r, column=_col('O'))
        c.value = f'=M{r}*(1-N{r})'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Margen Canal (de Info)
        c = ws.cell(row=r, column=_col('P'))
        c.value = '=Info!$J$10'
        _v(c, sz=8, bg=GREEN_LINK, color='166534'); _pct(c)

        # V/unit Venta = R/L
        c = ws.cell(row=r, column=_col('Q'))
        c.value = f'=R{r}/L{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Costo Canal = O*(1-P)
        c = ws.cell(row=r, column=_col('R'))
        c.value = f'=O{r}*(1-P{r})'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Costo Ingram Unit = U/L
        c = ws.cell(row=r, column=_col('S'))
        c.value = f'=U{r}/L{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Precio Compra Ingram = O*(1-MargenCadena)
        c = ws.cell(row=r, column=_col('T'))
        c.value = f'=O{r}*(1-Info!$J$9)'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Costo Ingram = O*0.95  (margen Ingram embebido)
        c = ws.cell(row=r, column=_col('U'))
        c.value = f'=O{r}*(1-Info!$J$11)'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Margen Ingram TX = 1-U/R
        c = ws.cell(row=r, column=_col('V'))
        c.value = f'=IF(R{r}<>0,1-U{r}/R{r},"")'
        _v(c, sz=8, bg=bg, color='000000'); _pct(c)

        # Precio unitario cliente final = O/L
        c = ws.cell(row=r, column=_col('X'))
        c.value = f'=O{r}/L{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # % Incentivos Total = F12+G12+F14+G14 (desde Info)
        c = ws.cell(row=r, column=_col('Z'))
        c.value = '=Info!$F$12+Info!$G$12+Info!$F$14+Info!$G$14'
        _v(c, sz=8, bg=GREEN_LINK, color='166534'); _pct(c)

        # Incentivo $ = Z*O
        c = ws.cell(row=r, column=_col('AA'))
        c.value = f'=Z{r}*O{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # % BP = D+E (CVR OnTime BP + CVR Renew BP)
        c = ws.cell(row=r, column=_col('AB'))
        c.value = f'=D{r}+E{r}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534'); _pct(c)

        # Total Incentivos BP = AB*O
        c = ws.cell(row=r, column=_col('AC'))
        c.value = f'=AB{r}*O{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # OnTime BP = O*D
        c = ws.cell(row=r, column=_col('AD'))
        c.value = f'=O{r}*D{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Renew BP = O*E
        c = ws.cell(row=r, column=_col('AE'))
        c.value = f'=O{r}*E{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Incentivo Total Unitario = AA/L
        c = ws.cell(row=r, column=_col('AF'))
        c.value = f'=IF(L{r}<>0,AA{r}/L{r},"")'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Incentivo BP Unitario = AC/L
        c = ws.cell(row=r, column=_col('AG'))
        c.value = f'=IF(L{r}<>0,AC{r}/L{r},"")'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Incentivo Ingram Unitario = (AA-AC)/L
        c = ws.cell(row=r, column=_col('AH'))
        c.value = f'=IF(L{r}<>0,(AA{r}-AC{r})/L{r},"")'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Costo Total BP = R-AC
        c = ws.cell(row=r, column=_col('AJ'))
        c.value = f'=R{r}-AC{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Costo Ingram Total = U-AA
        c = ws.cell(row=r, column=_col('AK'))
        c.value = f'=U{r}-AA{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Margen antes inc = 1-(AK_total/R_total) — en fila individual es ratio
        c = ws.cell(row=r, column=_col('AL'))
        tot_r = first_data_row + n  # fila de totales
        c.value = f'=IF(R{tot_r}<>0,1-(AK{tot_r}/R{tot_r}),"")'
        _v(c, sz=8, bg=bg, color='000000'); _pct(c)

        c = ws.cell(row=r, column=_col('AM'))
        c.value = f'=IF(AJ{tot_r}<>0,1-(AK{tot_r}/AJ{tot_r}),"")'
        _v(c, sz=8, bg=bg, color='000000'); _pct(c)

        ws.row_dimensions[r].height = 16

    # ── Fila de totales ──────────────────────────────────────────────────────
    r_tot = first_data_row + n
    ws.cell(row=r_tot, column=_col('L')).value = 'Total'
    _v(ws.cell(row=r_tot, column=_col('L')), bold=True, sz=9, bg=INGRAM_DARK, color=WHITE, align='center')

    for col in ['O','R','T','U','AA','AC','AD','AE','AJ','AK']:
        c = ws.cell(row=r_tot, column=_col(col))
        c.value = f'=SUM({col}{first_data_row}:{col}{r_tot-1})'
        _v(c, bold=True, sz=9, bg=INGRAM_DARK, color=WHITE)
        _usd(c)

    for col in ['V']:
        c = ws.cell(row=r_tot, column=_col(col))
        c.value = f'=IF(R{r_tot}<>0,1-U{r_tot}/R{r_tot},"")'
        _v(c, bold=True, sz=9, bg=INGRAM_DARK, color=WHITE)
        _pct(c)

    ws.row_dimensions[r_tot].height = 20

    # Nota de descuento (siguiente fila)
    r_nd = r_tot + 1
    ws.cell(row=r_nd, column=_col('L')).value = 'Nota de descuento'
    _v(ws.cell(row=r_nd, column=_col('L')), bold=True, sz=9)
    c = ws.cell(row=r_nd, column=_col('O'))
    c.value = f'=AC{r_tot}'
    _usd(c)
    _v(c, bold=True, sz=9, bg=YELLOW_INPUT)
    ws.row_dimensions[r_nd].height = 16

    return first_data_row, n, r_tot


# ── HOJA 3: COTIZACIÓN ────────────────────────────────────────────────────────
def _build_cotizacion(ws, lines, quote_number, bp_margin, first_data_row, n, r_tot):
    first = lines[0]

    for col, w in [('A',3),('B',20),('C',28),('D',20),('E',10),
                   ('F',16),('G',16),('H',10),('I',14),('J',16)]:
        ws.column_dimensions[col].width = w

    # ── Encabezado ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    ws['A1'] = 'INGRAM MICRO Colombia  ·  IBM Authorized Reseller'
    _h(ws['A1'], bg=INGRAM_DARK, sz=12)
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    ws['A2'] = 'IBM Software Subscription & Support (S&S) — Cotización al Canal'
    _h(ws['A2'], bg=INGRAM_BLUE, sz=10)
    ws.row_dimensions[2].height = 22

    # Datos del canal (jalados de Info)
    meta = [
        ('Canal:',    'C2', 'B3'),
        ('Cliente Final:', 'C3', 'B4'),  # no existen aún en Info como ref directa
        ('Fabricante:', 'IBM - International Business Machine', 'B5'),
        ('Fecha:',    None, 'B6'),
    ]
    ws['B3'] = 'Canal:'
    ws['C3'] = '=Info!D4'
    _v(ws['C3'], bg=GREEN_LINK, color='166534', sz=9)
    ws['B4'] = 'Cliente Final:'
    ws['C4'] = '=Info!D5'
    _v(ws['C4'], bg=GREEN_LINK, color='166534', sz=9)
    ws['B5'] = 'Fabricante:'
    ws['C5'] = 'IBM - International Business Machine'
    _v(ws['C5'], sz=9)
    ws['B6'] = 'Fecha:'
    ws['C6'] = '=TODAY()'
    _v(ws['C6'], sz=9); ws['C6'].number_format = DATE_FMT

    for row in range(3, 8):
        ws.row_dimensions[row].height = 16

    for cell in [ws['B3'],ws['B4'],ws['B5'],ws['B6']]:
        _v(cell, bold=True, color=GRAY_TEXT, sz=9)

    # Texto descripción
    ws.merge_cells('B8:J8')
    ws['B8'] = '="A continuación, se comparten precios aprobados bajo SBID  "&Info!J4&" para el cliente "&C4'
    _v(ws['B8'], sz=9)
    ws.row_dimensions[8].height = 16

    # ── Sección RESUMEN PROPUESTA ────────────────────────────────────────────
    ws.merge_cells('B10:J10')
    ws['B10'] = 'RESUMEN PROPUESTA ECONÓMICA'
    _h(ws['B10'], bg=INGRAM_DARK, sz=10, align='left')
    ws['B10'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[10].height = 20

    # Headers tabla cotización
    cot_headers = ['Fecha de Inicio','Fecha de Fin','Parte Número','Descripción',
                   'Cantidad','Precio Cliente\nFinal (MEP)','Margen Canal','V/unit Canal','Costo Total Canal']
    cot_cols = list('BCDEFGHIJ')
    for col, hdr in zip(cot_cols, cot_headers):
        c = ws[f'{col}11']
        c.value = hdr
        _h(c, bg=INGRAM_BLUE, sz=8, wrap=True)
    ws.row_dimensions[11].height = 28

    # Filas de datos (jaladas de Calculadora)
    # Calculadora: F=inicio, G=fin, H=NP, I=desc, L=qty, O=precio, P=margen, Q=v/unit, R=costo
    calc_map = {
        'B': 'F', 'C': 'G', 'D': 'H', 'E': 'I',
        'F': 'L', 'G': 'O', 'H': 'P', 'I': 'Q', 'J': 'R'
    }
    for i in range(n):
        r_cot = 12 + i
        calc_row = first_data_row + i
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        for cot_col, calc_col in calc_map.items():
            c = ws.cell(row=r_cot, column=_col(cot_col))
            c.value = f'=Calculadora!{calc_col}{calc_row}'
            _v(c, sz=8, bg=GREEN_LINK, color='166534')
            # Formato por columna
            if cot_col in ['G','I','J']:  _usd(c)
            if cot_col == 'H':            _pct(c)
            if cot_col in ['B','C']:      c.number_format = DATE_FMT
        ws.row_dimensions[r_cot].height = 16

    # Fila TOTAL
    r_ctot = 12 + n
    ws.merge_cells(f'B{r_ctot}:E{r_ctot}')
    ws[f'B{r_ctot}'] = 'TOTAL:'
    _h(ws[f'B{r_ctot}'], bg=INGRAM_DARK, sz=9)
    ws[f'G{r_ctot}'] = f'=SUM(G12:G{r_ctot-1})'
    _v(ws[f'G{r_ctot}'], bold=True, sz=9, bg=INGRAM_DARK, color=WHITE); _usd(ws[f'G{r_ctot}'])
    ws[f'J{r_ctot}'] = f'=SUM(J12:J{r_ctot-1})'
    _v(ws[f'J{r_ctot}'], bold=True, sz=9, bg=INGRAM_DARK, color=WHITE); _usd(ws[f'J{r_ctot}'])
    ws.row_dimensions[r_ctot].height = 20

    # Nota de descuento
    r_nd = r_ctot + 1
    ws[f'B{r_nd}'] = 'Nota de descuento'
    _v(ws[f'B{r_nd}'], bold=True, sz=9)
    ws[f'J{r_nd}'] = f'=Calculadora!O{r_tot+1}'
    _v(ws[f'J{r_nd}'], bold=True, sz=9, bg=YELLOW_INPUT); _usd(ws[f'J{r_nd}'])
    ws.row_dimensions[r_nd].height = 16

    # Texto legal
    r_leg = r_nd + 2
    ws.merge_cells(f'B{r_leg}:J{r_leg+1}')
    ws[f'B{r_leg}'] = (
        'Los precios aquí expresados son los precios a los que INGRAM MICRO SAS facturará al Canal. '
        'El precio al Cliente debe ser definido por el Canal o por el fabricante en su propuesta según las condiciones del negocio.'
    )
    ws[f'B{r_leg}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    _v(ws[f'B{r_leg}'], sz=8, color=GRAY_TEXT)
    ws.row_dimensions[r_leg].height = 28

    # Margen transaccional resumen
    r_mg = r_leg + 3
    ws.merge_cells(f'B{r_mg}:J{r_mg}')
    ws[f'B{r_mg}'] = f'Margen transaccional aproximado incluyendo nota de descuento: {int((bp_margin + 4)*100)/100}%  Detalle de incentivos:'
    _v(ws[f'B{r_mg}'], bold=True, sz=9)
    ws.row_dimensions[r_mg].height = 16

    # Tabla incentivos
    r_inc = r_mg + 1
    inc_headers = ['','Parte Número','Descripción','Cantidad','Total Incentivo','On time','Renew']
    inc_cols = list('BCDEFGH')
    for col, hdr in zip(inc_cols, inc_headers):
        c = ws[f'{col}{r_inc}']
        c.value = hdr
        _h(c, bg=INGRAM_BLUE, sz=8)
    ws.row_dimensions[r_inc].height = 18

    # Calculadora: H=NP, I=desc, L=qty, AC=total_inc, AD=ontime, AE=renew
    inc_map = {'C':'H','D':'I','E':'L','F':'AC','G':'AD','H':'AE'}
    for i in range(n):
        r_ii = r_inc + 1 + i
        calc_row = first_data_row + i
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        for cot_col, calc_col in inc_map.items():
            c = ws.cell(row=r_ii, column=_col(cot_col))
            c.value = f'=Calculadora!{calc_col}{calc_row}'
            _v(c, sz=8, bg=GREEN_LINK, color='166534')
            if cot_col in ['F','G','H']: _usd(c)
        ws.row_dimensions[r_ii].height = 16

    r_itot = r_inc + 1 + n
    ws[f'C{r_itot}'] = 'Total'
    _v(ws[f'C{r_itot}'], bold=True, sz=9, bg=INGRAM_DARK, color=WHITE)
    for col in ['F','G','H']:
        c = ws[f'{col}{r_itot}']
        c.value = f'=SUM({col}{r_inc+1}:{col}{r_itot-1})'
        _v(c, bold=True, sz=9, bg=INGRAM_DARK, color=WHITE); _usd(c)
    ws.row_dimensions[r_itot].height = 18

    # ── Condiciones generales ────────────────────────────────────────────────
    r_cond = r_itot + 2
    condiciones = [
        'Incentivos sujetos a cambio de programa Partner Plus de IBM',
        'Los precios enviados en esta propuesta están aprobados por IBM bajo el SBID',
        'El manejo de los SBID está regulado por lo establecido en el BPA firmado con IBM.',
        'Para procesar este SBID, Ingram Micro podrá solicitar la Orden de Compra (OC) enviada por el cliente final.',
        'Las notas de descuento se generan de manera independiente tal y como las realiza el fabricante.',
        'Propuesta válida por 30 días. Los precios están expresados en USD y sujetos a cambios según las políticas de IBM.',
        'Se respetará el Precio Máximo para el Usuario Final (MEP), según lo indicado por IBM y el distribuidor autorizado.',
        'Los valores aquí mencionados no incluyen IVA, dicho impuesto será cargado en la factura.',
    ]
    ws.merge_cells(f'B{r_cond}:J{r_cond}')
    ws[f'B{r_cond}'] = 'Condiciones Generales de Cotización – INGRAM MICRO S.A.S.'
    _h(ws[f'B{r_cond}'], bg=INGRAM_DARK, sz=9, align='left')
    ws[f'B{r_cond}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[r_cond].height = 18

    for j, cond in enumerate(condiciones):
        rr = r_cond + 1 + j
        ws.merge_cells(f'B{rr}:J{rr}')
        ws[f'B{rr}'] = f'• {cond}'
        _v(ws[f'B{rr}'], sz=8, color=GRAY_TEXT)
        ws.row_dimensions[rr].height = 14

    # Firma
    r_firma = r_cond + len(condiciones) + 2
    ws.merge_cells(f'B{r_firma}:J{r_firma}')
    ws[f'B{r_firma}'] = 'Angela Herrera · Business Development Manager IBM · angela.herrera@ingrammicro.com · +57 316 696 91 38'
    _v(ws[f'B{r_firma}'], bold=True, sz=8, color=INGRAM_BLUE)
    ws.row_dimensions[r_firma].height = 16


# ── HOJA 4: COMPRA ────────────────────────────────────────────────────────────
def _build_compra(ws, lines, quote_number, first_data_row, n):
    first = lines[0]

    widths = {
        'A':12,'B':12,'C':12,'D':22,'E':6,'F':16,'G':22,'H':10,
        'I':14,'J':10,'K':10,'L':12,'M':16,'N':6,'O':6,'P':10,
        'Q':10,'R':30,'S':14,'T':20,'U':3,'V':3,'W':3,
        'X':22,'Y':22,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Título ──────────────────────────────────────────────────────────────
    ws.merge_cells('A1:Y1')
    ws['A1'] = 'ORDEN DE COMPRA INTERNA — IBM S&S Renewals · Ingram Micro'
    _h(ws['A1'], bg=INGRAM_DARK, sz=11)
    ws.row_dimensions[1].height = 24

    # ── Headers ─────────────────────────────────────────────────────────────
    compra_headers = {
        'A':'TIPO DE SOLICITUD','B':'Código Canal','C':'Material','D':'IVA',
        'E':'Número de parte','F':'Descripción','G':'Cantidad',
        'H':'Precio de COMPRA UNITARIO','I':'Valor Acop 1 unitario',
        'J':'Valor Acop 2 unitario','K':'Valor Pocc',
        'L':'Nro. de Acop/Pocc,\nQuote, Deal, Soup, OPG',
        'M':'Nota descuento\nSi / No','N':'SKU','O':'Comentarios',
        'P':'Margen','Q':'Ship To','R':'Precio de Venta\nUNITARIO',
        'S':'TERMINO DE PAGO','T':'FIORI',
    }
    for col, hdr in compra_headers.items():
        c = ws[f'{col}2']
        c.value = hdr
        _h(c, bg=INGRAM_BLUE, sz=8, wrap=True)
    ws.row_dimensions[2].height = 30

    # Bloque info derecha
    info_right = [
        ('X2', '=Info!D5',     'Nombre del Cliente:'),
        ('X4', '=Info!D4',     'End User Segmentation:'),
        ('X5', 'CO-LED – Select Territory', None),
        ('X6', f'=Info!D4',    'Nombre del Reseller:'),
        ('X7', '=Info!G5',     'Site Number:'),
        ('X8', 'Renew',        'Tipo de Licencia:'),
        ('X9', f'=Info!J4',    'Número de SBID:'),
        ('X10','',             'Quote Number:'),
    ]
    for cell_addr, val, label in info_right:
        row = int(''.join(filter(str.isdigit, cell_addr)))
        if label:
            lc = ws.cell(row=row, column=_col('W'), value=label)
            _v(lc, bold=True, sz=8, color=GRAY_TEXT)
        c = ws[cell_addr]
        c.value = val
        _v(c, sz=8, bg=GREEN_LINK if str(val).startswith('=') else BLUE_INPUT)
        ws.row_dimensions[row].height = 16

    # ── Filas de datos ────────────────────────────────────────────────────────
    # Columnas Calculadora: H=NP, L=qty, Q=v/unit venta, S=costo ingram unit
    for i in range(n):
        r = 3 + i
        calc_row = first_data_row + i
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE

        _set(ws, r, 'A', 'Compra', bg, sz=8)

        # Código canal de Info
        c = ws.cell(row=r, column=_col('B'))
        c.value = '=Info!$D$6'
        _v(c, sz=8, bg=GREEN_LINK, color='166534')

        # NP desde Calculadora
        c = ws.cell(row=r, column=_col('F'))
        c.value = f'=Calculadora!H{calc_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534')

        # Cantidad
        c = ws.cell(row=r, column=_col('G'))
        c.value = f'=Calculadora!L{calc_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534')

        # Precio de compra unitario = costo ingram unit
        c = ws.cell(row=r, column=_col('H'))
        c.value = f'=Calculadora!S{calc_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534'); _usd(c)

        # Valor Acop 1/2 = N/A
        for col in ['I','J']:
            _set(ws, r, col, 'N/A', bg, sz=8, align='center')

        # Valor Pocc = H*G
        c = ws.cell(row=r, column=_col('K'))
        c.value = f'=H{r}*G{r}'
        _v(c, sz=8, bg=bg, color='000000'); _usd(c)

        # Nro SBID
        c = ws.cell(row=r, column=_col('L'))
        c.value = '=Info!$J$4'
        _v(c, sz=8, bg=GREEN_LINK, color='166534')

        # Nota descuento Si
        _set(ws, r, 'M', 'Si', bg, sz=8, align='center')

        # SKU N/A
        _set(ws, r, 'N', 'N/A', bg, sz=8, align='center')

        # Comentarios
        _set(ws, r, 'O', 'El margen real de la NV es el 2,06%, Se factura fecha de emisión de la factura', bg, sz=8)

        # Margen = 1-(H/S)
        c = ws.cell(row=r, column=_col('P'))
        c.value = f'=IF(S{r}<>0,1-(H{r}/S{r}),"")'
        _v(c, sz=8, bg=bg, color='000000'); _pct(c)

        # Ship To
        _set(ws, r, 'Q', 'Empresa, contacto celular, dirección, ciudad, etc.', bg, sz=8)

        # Precio de venta unitario = V/unit venta del canal
        c = ws.cell(row=r, column=_col('R'))
        c.value = f'=Calculadora!Q{calc_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534'); _usd(c)

        # Precio de venta: S = R (alias)
        c = ws.cell(row=r, column=_col('S'))
        c.value = f'=Calculadora!Q{calc_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534')

        _set(ws, r, 'T', 'Estándar del canal', bg, sz=8)

        ws.row_dimensions[r].height = 16

    # ── Tabla resumen para importar a Xvantage ────────────────────────────────
    r_xv = 3 + n + 2
    ws.merge_cells(f'B{r_xv}:G{r_xv}')
    ws[f'B{r_xv}'] = 'Resumen para importar a Xvantage'
    _h(ws[f'B{r_xv}'], bg=INGRAM_DARK, sz=9, align='left')
    ws[f'B{r_xv}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[r_xv].height = 18

    xv_headers = ['Material','Cant.','Venta Unit U$','Costo Unit U$','Importar Xvantage','Quote X4A']
    for j, hdr in enumerate(xv_headers):
        c = ws.cell(row=r_xv+1, column=_col('B')+j)
        c.value = hdr
        _h(c, bg=INGRAM_BLUE, sz=8)
    ws.row_dimensions[r_xv+1].height = 18

    for i in range(n):
        r_xi = r_xv + 2 + i
        src_row = 3 + i
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        # Material = NP
        c = ws.cell(row=r_xi, column=_col('B'))
        c.value = f'=F{src_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534')
        # Cantidad
        c = ws.cell(row=r_xi, column=_col('C'))
        c.value = f'=G{src_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534')
        # Venta Unit
        c = ws.cell(row=r_xi, column=_col('D'))
        c.value = f'=R{src_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534'); _usd(c)
        # Costo Unit
        c = ws.cell(row=r_xi, column=_col('E'))
        c.value = f'=H{src_row}'
        _v(c, sz=8, bg=GREEN_LINK, color='166534'); _usd(c)
        # Importar Xvantage (concatenado)
        c = ws.cell(row=r_xi, column=_col('F'))
        c.value = (
            f'=C{r_xi}&","&D{r_xi}&","'
            f'&TEXT(ROUNDUP(E{r_xi},2),"0.00")'
            f'&",,,,"'
            f'&TEXT(ROUNDUP(E{r_xi},2),"0.00")'
        )
        _v(c, sz=8, bg=bg)
        ws.row_dimensions[r_xi].height = 16


# ── Helpers internos ──────────────────────────────────────────────────────────
from openpyxl.utils import column_index_from_string

def _col(letter):
    return column_index_from_string(letter)

def _set(ws, row, col_letter, value, bg=WHITE, sz=9, align='left', fmt=None, wrap=False):
    c = ws.cell(row=row, column=_col(col_letter), value=value)
    _v(c, sz=sz, bg=bg, align=align, wrap=wrap)
    if fmt: c.number_format = fmt
    return c

# ── FUNCIÓN PRINCIPAL ─────────────────────────────────────────────────────────
def generate_internal_excel(lines, quote_number, output_path,
                            bp_margin=3.0, ingram_margin=2.1,
                            cvr_renew_bp=2.0, cvr_renew_ingram=2.0,
                            cvr_ontime_bp=2.0, cvr_ontime_ingram=1.0):
    """
    Genera el Excel de renovación IBM S&S con 4 hojas vinculadas por fórmulas.
    Replica la estructura de PLANTILLA_RNW.xlsx.

    Parámetros:
        lines          — lista de dicts del XML parser (con datos del XML)
        quote_number   — número de quote IBM
        output_path    — ruta de salida del .xlsx
        bp_margin      — margen canal BP en % (ej: 3.0)
        ingram_margin  — margen Ingram en % (ej: 2.1)
        cvr_renew_bp/ingram  — CVR Renew en %
        cvr_ontime_bp/ingram — CVR OnTime en %
    """
    # Extraer márgenes desde las líneas si ya fueron calculados por _apply_margins
    if lines and 'bp_margin' in lines[0]:
        bp_margin      = lines[0].get('bp_margin',       bp_margin)
        ingram_margin  = lines[0].get('ingram_margin',   ingram_margin)
        cvr_renew_bp   = lines[0].get('cvr_renew_bp',    cvr_renew_bp)
        cvr_ontime_bp  = lines[0].get('cvr_ontime_bp',   cvr_ontime_bp)
        # Estos dos no los guardaba _apply_margins — los tomamos del parámetro
        # (ya vienen correctos desde _build_quote_files con el Bug 1 corregido)

    wb = Workbook()

    # ── Hoja 1: Info ────────────────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = 'Info'
    _build_info(ws_info, lines, quote_number, bp_margin, ingram_margin,
                cvr_renew_bp, cvr_renew_ingram, cvr_ontime_bp, cvr_ontime_ingram)

    # ── Hoja 2: Calculadora ─────────────────────────────────────────────────
    ws_calc = wb.create_sheet('Calculadora')
    first_data_row, n, r_tot = _build_calculadora(ws_calc, lines)

    # ── Hoja 3: Cotización ──────────────────────────────────────────────────
    ws_cot = wb.create_sheet('Cotización')
    _build_cotizacion(ws_cot, lines, quote_number, bp_margin, first_data_row, n, r_tot)

    # ── Hoja 4: Compra ──────────────────────────────────────────────────────
    ws_comp = wb.create_sheet('Compra')
    _build_compra(ws_comp, lines, quote_number, first_data_row, n)

    wb.save(output_path)
    return output_path
